# This Python file uses the following encoding: utf-8
# controller/projects_controller.py

import config
import os
import zipfile
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from Dialog.project_dialog import ProjectDialog
from PySide6.QtWidgets import QFileDialog
from PySide6.QtCore import QObject, Signal, QDir, QFile, QFileInfo, QUrl
from PySide6.QtGui import QDesktopServices

class ProjectsController(QObject):
    """
    Controller for managing project operations in the application.

    This class acts as the intermediary between the project view and the data models. It handles user interactions, business logic, and updates the view based on changes in the model.

    Attributes
    ----------
    project_opened : Signal
        Signal emitted when a project is opened.
    project_saved : Signal
        Signal emitted when a project is saved.
    project_closed : Signal
        Signal emitted when a project is closed.

    Methods
    -------
    """
    project_opened = Signal(dict)
    project_saved = Signal(dict)
    project_closed = Signal()

    def __init__(self, projects_view, projects_model, actors_model, useCases_model,
                 technicalFactors_model, environmentalFactors_model, dashboard_model):
        """
        Initialize the ProjectsController.

        Parameters
        ----------
        projects_view : ProjectsView
            The view for projects.
        projects_model : ProjectsModel
            The model for projects.
        actors_model : ActorsModel
            The model for actors.
        useCases_model : UseCasesModel
            The model for use cases.
        technicalFactors_model : TechnicalFactorsModel
            The model for technical factors.
        environmentalFactors_model : EnvironmentalFactorsModel
            The model for environmental factors.
        dashboard_model : DashboardModel
            The model for the dashboard.
        """
        super().__init__()
        self.view = projects_view
        self.projects_model = projects_model
        self.actors_model = actors_model
        self.useCases_model = useCases_model
        self.technicalFactors_model = technicalFactors_model
        self.environmentalFactors_model = environmentalFactors_model
        self.dashboard_model = dashboard_model
        self.connect_signals()
        self.load_projects()

    def connect_signals(self):
        """
        Connect signals from the view to the appropriate handler methods.
        """
        self.view.text_changed.connect(self.filter_projects_table)
        self.view.project_managed.connect(self.open_management_project_dialog)
        self.view.project_imported.connect(self.import_project)
        self.view.project_deleted.connect(self.delete_project)
        self.view.project_favorite.connect(self.set_favorite_project)
        self.view.project_selected.connect(self.open_project)
        self.view.project_downloaded.connect(self.download_project)
        self.view.projects_exported.connect(self.export_projects)
        self.view.report_request.connect(self.generate_excel_report)

    def open_project(self, project_data, row):
        """
        Open a project.

        Parameters
        ----------
        project_data : dict
            Dictionary containing project data.
        row : int
            The row number in the table.
        """
        project_id = project_data['id']
        project_data['row'] = row
        last_access = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

        try:
            self.projects_model.start_transaction()

            # Load dashboard data
            return_value = self.dashboard_model.load_dashboard_data(project_id)
            if return_value != config.SUCCESS:
                raise Exception("Failed to load dashboard data")

            # Load actors data
            return_value = self.actors_model.load_actors_data(project_id)
            if return_value != config.SUCCESS:
                raise Exception("Failed to load actors data")

            # Load use cases data
            return_value = self.useCases_model.load_use_cases_data(project_id)
            if return_value != config.SUCCESS:
                raise Exception("Failed to load use cases data")

            # Load technical factors data
            return_value = self.technicalFactors_model.load_technical_factors_data(project_id)
            if return_value != config.SUCCESS:
                raise Exception("Failed to load technical factors data")

            # Load environmental factors data
            return_value = self.environmentalFactors_model.load_environmental_factors_data(project_id)
            if return_value != config.SUCCESS:
                raise Exception("Failed to load environmental factors data")

            # Update last access
            return_value = self.projects_model.update_last_access(project_id, last_access)
            if return_value != config.SUCCESS:
                raise Exception("Failed to update last access")

            # If all operations succeed
            self.projects_model.commit()
            self.view.update_last_access(row, last_access)
            self.project_opened.emit(project_data)

        except Exception:
            # Handle exceptions and rollback transaction
            self.projects_model.rollback()
            self.view.display_message("Failed Operation", "One or more components couldn't be loaded.", config.CRITICAL_IMG)

    def filter_projects_table(self, search_text):
        """
        Filter the projects table based on search text.

        Parameters
        ----------
        search_text : str
            The text to search for in the projects table.
        """
        rows_to_show = self.view.filter_table(search_text)
        self.view.update_table_visibility(rows_to_show)

    def open_management_project_dialog(self, action, data_send, project_row):
        """
        Open the project management dialog.

        Parameters
        ----------
        action : str
            The action to perform ("new" or "edit").
        data_send : dict
            Data to be sent to the dialog.
        project_row : int
            The row number in the table.
        """
        self.managementProject_Dialog = ProjectDialog(action, data_send)
        if action == "new":
           self.managementProject_Dialog.data_saved.connect(self.create_project)
        else:
            self.managementProject_Dialog.data_saved.connect(lambda data_saved: self.edit_project(data_send, data_saved, project_row))
        self.managementProject_Dialog.exec()

    def edit_project(self, data_send, data_saved, row):
        """
        Edit an existing project.

        Parameters
        ----------
        data_send : dict
            Original data of the project to be edited.
        data_saved : dict
            New data for the project.
        row : int
            The row number in the table.
        """
        project_id = data_send['id']
        result = self.projects_model.update_project(project_id, data_saved)
        if result == config.SUCCESS:
            self.managementProject_Dialog.accept()
            self.view.update_projects_table(data_saved, row)
            data_saved['row'] = row
            data_saved['id'] = data_send['id']
            self.project_saved.emit(data_saved)
            self.view.display_message("Successful Operation", "The operation was completed successfully.", config.INFORMATION_IMG)
        elif result == config.ALREADY_EXIST:
            self.view.display_message("Warning", "There is already a project with that name. Please, try again.", config.WARNING_IMG)
        elif result == config.NOT_EXIST:
            self.managementProject_Dialog.accept()
            self.view.display_message("Warning", "The project does not exist.", config.WARNING_IMG)
        elif result == config.FAILURE:
            self.managementProject_Dialog.accept()
            self.view.display_message("Failed Operation", "The operation could not be completed.", config.CRITICAL_IMG)

    def create_project(self, data_saved, external_transaction=False):
        """
        Create a new project.

        Parameters
        ----------
        data_saved : dict
            Data for the new project.
        external_transaction : bool, optional
            Indicates if the transaction is external, by default False.

        Returns
        -------
        tuple
            Status code and project ID.
        """
        try:
            if not external_transaction:
                self.projects_model.start_transaction()

            result, project_id = self.projects_model.add_project(data_saved)
            if result == config.SUCCESS:
                if not external_transaction:
                    self.managementProject_Dialog.accept()
                if self.projects_model.insert_parameters(project_id) != config.SUCCESS:
                    raise Exception("Failed to add project. Parameters")
                if self.technicalFactors_model.insert_technical_factors_default(project_id) != config.SUCCESS:
                    raise Exception("Failed to add project. TF")
                if self.environmentalFactors_model.insert_environmental_factors_default(project_id) != config.SUCCESS:
                    raise Exception("Failed to add project. EF")

                if not external_transaction:
                    self.projects_model.commit()

                if not external_transaction:
                    data_saved['id'] = str(project_id)
                    self.view.update_projects_table(data_saved)
                    self.view.display_message(
                        "Successful Operation", "The operation was completed successfully.",
                        config.INFORMATION_IMG, ok_callback=lambda: self.view.open_project(project_id=project_id)
                    )

                return config.SUCCESS, project_id

            elif result == config.ALREADY_EXIST:
                raise Exception("Project already exists")
            elif result == config.TOO_MANY_PROJECTS:
                if not external_transaction:
                    self.managementProject_Dialog.accept()
                raise Exception("Reached the maximum number of allowed projects")
            elif result == config.FAILURE:
                if not external_transaction:
                    self.managementProject_Dialog.accept()
                raise Exception("Failed to add project")
        except Exception as e:
            if not external_transaction:
                self.projects_model.rollback()
            if "Project already exists" in str(e):
                self.view.display_message("Warning", "There is already a project with that name. Please, try again.", config.WARNING_IMG)
            elif "Reached the maximum number of allowed projects" in str(e):
                self.view.display_message("Warning", f"You have reached the maximum number of allowed projects ({config.PROJECT_LIMIT}). Please delete an existing project before adding a new one.", config.WARNING_IMG)
            else:
                self.view.display_message("Failed Operation", "The operation could not be completed.", config.CRITICAL_IMG)

    def delete_project(self, project_id, table_row):
        """
        Delete a project.

        Parameters
        ----------
        project_id : int
            The ID of the project to be deleted.
        table_row : int
            The row number in the table.
        """
        reply = self.view.display_message("Delete confirmation", "Are you sure you want to delete the selected project?", config.CONFIRMATION_IMG, "confirmation_message")
        if reply:
            result = self.projects_model.delete_project(project_id)
            if result == config.SUCCESS:
                self.view.remove_table_row(table_row)
                self.close_project(confirmation=False)
                self.view.display_message("Successful Operation", "The project was deleted successfully.", config.INFORMATION_IMG)
            elif result == config.NOT_EXIST:
                self.view.display_message("Warning", "The project to be deleted does not exist.", config.WARNING_IMG)
            elif result == config.FAILURE:
                self.view.display_message("Failed Operation", "The operation could not be completed.", config.CRITICAL_IMG)

    def export_projects(self):
        """
        Export all projects to a ZIP file.
        """
        zip_file_path, _  = QFileDialog.getSaveFileName(None, "Save QuickEst projects", "quickest_projects.zip", "ZIP files (*.zip);; All files (*)")

        if not zip_file_path:
            return

        if not zip_file_path.endswith('.zip'):
            self.view.display_message("Warning", "The operation could not be completed: missing extension '.zip'.", config.WARNING_IMG)
            return

        return_value, projects = self.projects_model.get_projects()
        if return_value != config.SUCCESS:
            self.view.display_message("Failed Operation", "Projects couldn't be exported.", config.CRITICAL_IMG)
            return

        temp_directory = os.path.join(os.path.dirname(zip_file_path), "quickest_temp_export_directory")
        if not os.path.exists(temp_directory):
            os.makedirs(temp_directory)

        self.projects_model.start_transaction()
        project_directories = []

        try:
            for project in projects:
                project_directory = self.download_project(project['id'], project['name'], temp_directory, transaction_active=False)
                if project_directory == config.FAILURE:
                    raise Exception(f"Failed to export project {project['name']}")
                project_directories.append(project_directory)

            with zipfile.ZipFile(zip_file_path, 'w') as zip_file:
                for project_directory in project_directories:
                    for folder_name, subfolders, filenames in os.walk(project_directory):
                        for filename in filenames:
                            file_path = os.path.join(folder_name, filename)
                            zip_file.write(file_path, os.path.relpath(file_path, temp_directory))

            self.projects_model.commit()

            for project_directory in project_directories:
                if QDir(project_directory).exists():
                    self.delete_directory(project_directory)

            if os.path.exists(temp_directory):
                self.delete_directory(temp_directory)

            self.view.display_message("Successful Operation", f"All projects have been successfully exported to {zip_file_path}.", config.INFORMATION_IMG)

        except Exception:
            self.projects_model.rollback()

            for project_directory in project_directories:
                if QDir(project_directory).exists():
                    self.delete_directory(project_directory)

            if zip_file_path and os.path.exists(zip_file_path):
                os.remove(zip_file_path)

            if temp_directory and os.path.exists(temp_directory):
                self.delete_directory(temp_directory)

            self.view.display_message("Failed Operation", "The operation could not be completed.", config.CRITICAL_IMG)

    def download_project(self, project_id, project_name, directory=None, transaction_active=True):
        """
        Download a project to a specified directory.

        Parameters
        ----------
        project_id : int
            The ID of the project to download.
        project_name : str
            The name of the project.
        directory : str, optional
            The directory to save the project, by default None.
        transaction_active : bool, optional
            Indicates if the transaction is active, by default True.

        Returns
        -------
        str or int
            The directory where the project is downloaded or a status code indicating failure.
        """
        if directory is None:
            file_path, _  = QFileDialog.getSaveFileName(None, "Save Project", project_name, "All files (*)")
            if not file_path:
                return config.FAILURE

            directory = os.path.dirname(file_path)

        project_directory = QDir(directory).filePath(project_name)
        QDir().mkpath(project_directory)

        if transaction_active:
            self.projects_model.start_transaction()

        try:
            if self.export_project_data(project_id, project_name, project_directory) == config.FAILURE:
                raise Exception("Failed to export project data")

            if transaction_active:
                self.projects_model.commit()
                self.view.display_message("Successful Operation", f"The project '{project_name}' has been successfully exported to {project_directory}.", config.INFORMATION_IMG)
            return project_directory
        except Exception:
            if transaction_active:
                self.projects_model.rollback()
            if QDir(project_directory).exists():
                self.delete_directory(project_directory)
            if transaction_active:
                self.view.display_message("Failed Operation", "The operation could not be completed.", config.CRITICAL_IMG)
            return config.FAILURE

    def export_project_data(self, project_id, project_name, project_directory):
        """
        Export project data to the specified directory.

        Parameters
        ----------
        project_id : int
            The ID of the project to export.
        project_name : str
            The name of the project.
        project_directory : str
            The directory to save the project data.

        Returns
        -------
        int
            Status code indicating success or failure.
        """
        exports = [
            (self.projects_model, 'export_project_data', config.PROJECT_FILE, 'project'),
            (self.projects_model, 'export_parameters_data', config.PARAMETERS_FILE, 'parameters'),
            (self.actors_model, 'export_actors_data', config.ACTORS_FILE, 'actors'),
            (self.useCases_model, 'export_use_cases_data', config.USE_CASES_FILE, 'use_cases'),
            (self.technicalFactors_model, 'export_technical_factors_data', config.TECHNICAL_FACTORS_FILE, 'technical_factors'),
            (self.environmentalFactors_model, 'export_environmental_factors_data', config.ENVIRONMENTAL_FACTORS_FILE, 'environmental_factors')
        ]

        project_manifest = {}

        try:
            for model, method, file_name, table in exports:
                destination_file = QDir(project_directory).filePath(file_name)
                result = getattr(model, method)(project_id)
                if result != config.FAILURE:
                    if self.projects_model.write_data_to_binary_file(result, destination_file) == config.FAILURE:
                        raise Exception(f"Failed to write data to binary file: {file_name}")
                    else:
                        file_hash = self.projects_model.calculate_file_hash(destination_file)
                        project_manifest[table] = {"file_name": file_name, "hash": file_hash}
                else:
                    raise Exception(f"Failed to export data for {table}")
            file_path = f"{project_name}{config.PROJECT_FILE_EXTENSION}"
            if self.projects_model.write_project_manifest(file_path, project_directory, project_manifest) == config.FAILURE:
                raise Exception("Failed to write project manifest")

            return config.SUCCESS
        except Exception:
            return config.FAILURE

    def delete_directory(self, directory_path):
        """
        Delete a directory and its contents.

        Parameters
        ----------
        directory_path : str
            The path of the directory to delete.
        """
        dir = QDir(directory_path)
        for file_info in dir.entryInfoList(QDir.Files | QDir.Dirs | QDir.NoDotAndDotDot, QDir.NoSort):
            if file_info.isDir():
                self.delete_directory(file_info.absoluteFilePath())
            else:
                QFile.remove(file_info.absoluteFilePath())
        QDir().rmdir(directory_path)

    def set_favorite_project(self, project_id, table_row, is_favorite):
        """
        Set or unset a project as favorite.

        Parameters
        ----------
        project_id : int
            The ID of the project.
        table_row : int
            The row number in the table.
        is_favorite : bool
            Current favorite status of the project.
        """
        if is_favorite:
            result = self.projects_model.update_favorite_project(project_id, False)
        else:
            result = self.projects_model.update_favorite_project(project_id, True)
        if result == config.SUCCESS:
            self.view.update_favorite_button(table_row, is_favorite)
        elif result == config.NOT_EXIST:
            self.view.display_message("Warning", "The project does not exist.", config.WARNING_IMG)
        elif result == config.FAILURE:
            self.view.display_message("Failed Operation", "The operation could not be completed.", config.CRITICAL_IMG)

    def load_projects(self):
        """
        Load all projects.
        """
        return_value, projects = self.projects_model.get_projects()
        if return_value == config.SUCCESS:
            for project in projects:
                self.view.update_projects_table(project, None)
        elif return_value == config.FAILURE:
            self.view.display_message("Failed Operation", "Projects couldn't be loaded.", config.CRITICAL_IMG)

    def import_project(self):
        """
        Import a project from a directory.
        """
        directory = QFileDialog.getExistingDirectory(self.view, "Select directory", "", QFileDialog.ShowDirsOnly)
        if not directory:
            return

        manifest_file = self.projects_model.find_manifest(directory)
        if manifest_file == config.FAILURE:
            self.view.display_message("Warning", "No QuickEst project found in the selected directory. Please, select a valid project.", config.WARNING_IMG)
            return

        try:
            project_manifest = self.projects_model.load_json(manifest_file)
            if project_manifest == config.FAILURE:
                raise Exception("Failed to add project")

            loaded_data = self.projects_model.load_project_data(project_manifest, directory)
            if loaded_data == config.FAILURE:
                raise Exception("Failed to add project")
            return_value = self.add_project_to_database(directory, loaded_data)
            if return_value != config.FAILURE:
                project_data = return_value
                self.view.update_projects_table(project_data, None)
                self.view.display_message("Successful Operation", "The operation was completed successfully.", config.INFORMATION_IMG, ok_callback=lambda: self.view.open_project(project_id=project_data['id']))

        except Exception:
            self.view.display_message("Failed Operation", "No QuickEst project found in the selected directory. Please, select a valid project.", config.WARNING_IMG)

    def add_project_to_database(self, directory, loaded_data):
        """
        Add a project to the database.

        Parameters
        ----------
        directory : str
            The directory of the project.
        loaded_data : dict
            The loaded project data.

        Returns
        -------
        dict or int
            Project data dictionary or failure code.
        """
        try:
            self.projects_model.start_transaction()

            for project in loaded_data['project']:
                project_name = QFileInfo(directory).fileName().strip()
                if len(project_name) > 20:
                    raise ValueError("Project name cannot exceed 20 characters.")
                project_data = {
                    "favorite": 0,
                    "name": project_name,
                    "description": project.get('description'),
                    "created_at": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
                    "last_access": "––"
                }

            return_value, project_id = self.create_project(project_data, external_transaction=True)
            if return_value != config.SUCCESS:
                raise Exception("Failed to add project")

            if not self.insert_project_data(loaded_data, project_id):
                raise Exception("Failed to insert project data")

            self.projects_model.commit()
            project_data['id'] = project_id
            return project_data

        except ValueError as ve:
            self.view.display_message("Invalid Project Name", str(ve), config.WARNING_IMG)
            return config.FAILURE
        except Exception:
            self.projects_model.rollback()
            return config.FAILURE

    def insert_project_data(self, loaded_data, project_id):
        """
        Insert project data into the database.

        Parameters
        ----------
        loaded_data : dict
            The loaded project data.
        project_id : int
            The ID of the project.

        Returns
        -------
        bool
            True if successful, False otherwise.
        """
        try:
            for parameter in loaded_data["parameters"]:
                if self.dashboard_model.update_cf(
                    parameter.get('cf'),
                    project_id
                ) != config.SUCCESS:
                    return False

                percentages = {
                    'analysis': parameter.get('analysis_percentage'),
                    'design': parameter.get('design_percentage'),
                    'programming': parameter.get('programming_percentage'),
                    'testing': parameter.get('testing_percentage'),
                    'overloading': parameter.get('overloading_percentage')
                }
                if self.dashboard_model.update_percentages(percentages, project_id) != config.SUCCESS:
                    return False

                actors_weights = {
                    'simple_weight': parameter.get('actors_simple_weight'),
                    'average_weight': parameter.get('actors_average_weight'),
                    'complex_weight': parameter.get('actors_complex_weight')
                }
                if self.actors_model.update_actors_weights(actors_weights, project_id) != config.SUCCESS:
                    return False

                use_cases_weights = {
                    'simple_weight': parameter.get('use_cases_simple_weight'),
                    'average_weight': parameter.get('use_cases_average_weight'),
                    'complex_weight': parameter.get('use_cases_complex_weight')
                }
                if self.useCases_model.update_use_cases_weights(use_cases_weights,project_id) != config.SUCCESS:
                    return False

            for actor in loaded_data["actors"]:
                actor_data = {
                    'code': actor.get('code'),
                    'name': actor.get('name'),
                    'complexity': actor.get('complexity'),
                    'comment': actor.get('comment')
                }
                return_value, actor_id = self.actors_model.create_actor(actor_data, project_id)
                if return_value != config.SUCCESS:
                    return False

            for use_case in loaded_data["use_cases"]:
                use_case_data = {
                    'code': use_case.get('code'),
                    'name': use_case.get('name'),
                    'complexity': use_case.get('complexity'),
                    'transactions': use_case.get('transactions'),
                    'comment': use_case.get('comment')
                }
                return_value, useCase_id = self.useCases_model.create_use_case(use_case_data, project_id)
                if return_value != config.SUCCESS:
                    return False

            for t_factor in loaded_data["technical_factors"]:
                t_factor_data = {
                    'factor' :  t_factor.get('factor'),
                    'weight' :  t_factor.get('weight'),
                    'influence' :  t_factor.get('influence'),
                    'comment' :  t_factor.get('comment'),
                }
                if self.technicalFactors_model.update_technical_factor(t_factor_data, project_id) != config.SUCCESS:
                    return False

            for e_factor in loaded_data["environmental_factors"]:
                e_factor_data = {
                    'factor' :  e_factor.get('factor'),
                    'weight' :  e_factor.get('weight'),
                    'influence' :  e_factor.get('influence'),
                    'comment' :  e_factor.get('comment'),
                }
                if self.environmentalFactors_model.update_environmental_factor(e_factor_data, project_id) != config.SUCCESS:
                    return False

            return True

        except Exception:
            return False

    def generate_excel_report(self, project_id):
        """
        Generate an Excel report for a project.

        Parameters
        ----------
        project_id : int
            The ID of the project.
        """
        try:
            self.projects_model.start_transaction()
            effort_distribution_data = self.dashboard_model.get_effort_distribution_data(project_id)
            metrics_data = self.dashboard_model.get_metrics_data(project_id)
            project_data = self.projects_model.get_project_data(project_id)
            if project_data == config.FAILURE:
                raise Exception("Failed to retrieve project data")

            actors_data = self.actors_model.get_data(project_id)
            if actors_data == config.FAILURE:
                raise Exception("Failed to retrieve actors data")

            useCases_data = self.useCases_model.get_data(project_id)
            if useCases_data == config.FAILURE:
                raise Exception("Failed to retrieve use cases data")

            technicalFactors_data = self.technicalFactors_model.get_data(project_id)
            if technicalFactors_data == config.FAILURE:
                raise Exception("Failed to retrieve technical factors data")

            environmentalFactors_data = self.environmentalFactors_model.get_data(project_id)
            if environmentalFactors_data == config.FAILURE:
                raise Exception("Failed to retrieve environmental factors data")

            project_name = project_data['project_data']['Project Name'].iloc[0]
            file_name = f"{project_name}.xlsx"

            excel_file_path, _  = QFileDialog.getSaveFileName(None, "Save Project Report", file_name, "Excel files (*.xlsx);; All files (*)")

            if not excel_file_path:
                return

            if not excel_file_path.endswith('.xlsx'):
                raise Exception("missing extension '.xlsx'")

            with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
                project_data['project_data'].to_excel(writer, sheet_name='Results', index=False, startrow=1)
                project_data['report_date'].to_excel(writer, sheet_name='Results', index=False, startrow=4)
                metrics_data['total_effort'].to_excel(writer, sheet_name='Results', index=False, startrow=7)
                sheet = writer.sheets['Results']
                sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
                sheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
                sheet.merge_cells(start_row=8, start_column=1, end_row=8, end_column=2)
                sheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=2)
                max_column = project_data['project_data'].shape[1]
                self.add_formatted_title(writer.sheets['Results'], "RESULTS", 1, 1, max_column, "FFFF00")

                metrics_data['metrics_data'].to_excel(writer, sheet_name='Estimation Metrics', index=False, startrow=1)
                max_column = metrics_data['metrics_data'].shape[1]
                self.add_formatted_title(writer.sheets['Estimation Metrics'], "ESTIMATION METRICS", 1, 1, max_column, "FFFF00")

                effort_distribution_data.to_excel(writer, sheet_name='Effort Distribution', index=False, startrow=1)
                max_column = effort_distribution_data.shape[1]
                self.add_formatted_title(writer.sheets['Effort Distribution'], "EFFORT DISTRIBUTION", 1, 1, max_column, "FFFF00")

                actors_data['actors_data'].to_excel(writer, sheet_name='Actors', index=False, startrow=1)
                max_column_actors = actors_data['actors_data'].shape[1]
                self.add_formatted_title(writer.sheets['Actors'], "ACTORS DATA", 1, 1, max_column_actors, "FFCCFFFF")

                start_col_summary = max_column_actors + 1
                actors_data['actors_summary'].to_excel(writer, sheet_name='Actors', index=False, startcol=start_col_summary, startrow=1)
                max_column_summary = start_col_summary + actors_data['actors_summary'].shape[1]
                self.add_formatted_title(writer.sheets['Actors'], "ACTORS SUMMARY", 1, start_col_summary + 1, max_column_summary, "FFCCFFCC")

                useCases_data['useCases_data'].to_excel(writer, sheet_name='Use Cases', index=False, startrow=1)
                max_column_useCases = useCases_data['useCases_data'].shape[1]
                self.add_formatted_title(writer.sheets['Use Cases'], "USE CASES DATA", 1, 1, max_column_useCases, "FFCCFFFF")

                start_col_summary = max_column_useCases + 1
                useCases_data['useCases_summary'].to_excel(writer, sheet_name='Use Cases', index=False, startcol=start_col_summary, startrow=1)
                max_column_summary = start_col_summary + useCases_data['useCases_summary'].shape[1]
                self.add_formatted_title(writer.sheets['Use Cases'], "USE CASES SUMMARY", 1, start_col_summary + 1, max_column_summary, "FFCCFFCC")

                technicalFactors_data['technicalFactors_data'].to_excel(writer, sheet_name='Technical Factors', index=False, startrow=1)
                max_column_technicalFactors = technicalFactors_data['technicalFactors_data'].shape[1]
                self.add_formatted_title(writer.sheets['Technical Factors'], "TECHNICAL FACTORS DATA", 1, 1, max_column_technicalFactors, "FFCCFFFF")

                start_col_summary = max_column_technicalFactors + 1
                technicalFactors_data['technicalFactors_summary'].to_excel(writer, sheet_name='Technical Factors', index=False, startcol=start_col_summary, startrow=1)
                max_column_summary = start_col_summary + technicalFactors_data['technicalFactors_summary'].shape[1]
                self.add_formatted_title(writer.sheets['Technical Factors'], "TECHNICAL FACTORS SUMMARY", 1, start_col_summary + 1, max_column_summary, "FFCCFFCC")

                environmentalFactors_data['environmentalFactors_data'].to_excel(writer, sheet_name='Environmental Factors', index=False, startrow=1)
                max_column_environmentalFactors = environmentalFactors_data['environmentalFactors_data'].shape[1]
                self.add_formatted_title(writer.sheets['Environmental Factors'], "ENVIRONMENTAL FACTORS DATA", 1, 1, max_column_environmentalFactors, "FFCCFFFF")

                start_col_summary = max_column_environmentalFactors + 1
                environmentalFactors_data['environmentalFactors_summary'].to_excel(writer, sheet_name='Environmental Factors', index=False, startcol=start_col_summary, startrow=1)
                max_column_summary = start_col_summary + environmentalFactors_data['environmentalFactors_summary'].shape[1]
                self.add_formatted_title(writer.sheets['Environmental Factors'], "ENVIRONMENTAL FACTORS SUMMARY", 1, start_col_summary + 1, max_column_summary, "FFCCFFCC")

                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    max_column = len(worksheet['1'])
                    self.auto_adjust_column_width(worksheet, [(1, max_column)])

            self.projects_model.commit()
            success = QDesktopServices.openUrl(QUrl.fromLocalFile(excel_file_path))
            if not success:
                return

        except Exception as e:
            self.projects_model.rollback()
            if "missing extension" in str(e):
                self.view.display_message("Warning", f"The operation could not be completed: {e}. Please try again.", config.WARNING_IMG)
            else:
                self.view.display_message("Failed Operation", f"An error occurred while downloading the report: {e}. Please try again.", config.CRITICAL_IMG)

    def add_formatted_title(self, worksheet, title, start_row, start_column, end_column, bg_color):
        """
        Add a formatted title to an Excel sheet.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet to add the title to.
        title : str
            The text of the title.
        start_row : int
            The row where the title starts.
        start_column : int
            The column where the title starts.
        end_column : int
            The column where the title ends.
        bg_color : str
            Background color of the title in hexadecimal format.
        """
        title_cell = worksheet.cell(row=start_row, column=start_column)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=12)
        title_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        worksheet.merge_cells(start_row=start_row, start_column=start_column, end_row=start_row, end_column=end_column)

    def auto_adjust_column_width(self, worksheet, column_ranges, extra_width=4, max_width=50):
        """
        Automatically adjust the width of specified columns in an Excel worksheet based on their content.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet to adjust.
        column_ranges : list of tuple
            List of tuples specifying the start and end columns to adjust.
        extra_width : int, optional
            Additional width to add to the calculated column width, by default 4.
        max_width : int, optional
            Maximum column width, by default 50.
        """
        for start_col, end_col in column_ranges:
            for col_idx in range(start_col, end_col + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in worksheet[column_letter]:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, min(cell_length, max_width))
                adjusted_width = min(max_length + extra_width, max_width)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                for cell in worksheet[column_letter]:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    def close_project(self, confirmation=True):
        """
        Close the current project.

        Parameters
        ----------
        confirmation : bool, optional
            Whether to ask for confirmation before closing the project, by default True.
        """
        if confirmation:
            reply = self.view.display_message("Confirmation Message", "Are you sure you want to close the current project?", config.CONFIRMATION_IMG, "confirmation_message")
            if reply:
                self.project_closed.emit()
        else:
            self.project_closed.emit()
